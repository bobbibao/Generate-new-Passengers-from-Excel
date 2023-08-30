import io
import random
import datetime
import pandas as pd
from matplotlib import colors
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from flask import Flask, request, send_file, render_template
from flask.views import MethodView

app = Flask(__name__)

@app.get('/')
def index():
	return render_template('index.html')

class ExcelGeneratorView(MethodView):
    def convert_color_to_argb_hex(self, color):
        try:
            rgba = colors.to_rgba(color)
            argb_hex = f"{int(rgba[3] * 255):02X}{int(rgba[0] * 255):02X}{int(rgba[1] * 255):02X}{int(rgba[2] * 255):02X}"
            return argb_hex
        except ValueError:
            raise ValueError("Mã màu không tồn tại.")

    def generate_seat_numbers(self, num_passengers):
        available_seats = [f"{i}{j}" for i in range(10, int(num_passengers/6)+11) for j in ['A', 'B', 'C', 'D', 'E', 'G']]
        selected_seats = random.sample(available_seats, num_passengers)
        return selected_seats

    def generate_data(self, num_new_names, vietnamese_last_names, middle_names, first_names, airline_name, AUSTRALIAN_NAMES, FOREIGN_NAMES):
        num_vietnamese = int(num_new_names * 0.6) 
        num_australian = int(num_new_names * 0.3) 
        num_foreign = num_new_names - num_vietnamese - num_australian

        vietnamese_names = []
        australian_names = []
        foreign_names = []

        for _ in range(num_vietnamese):
            new_last_name = random.choice(vietnamese_last_names)
            new_middle_name = random.choice(middle_names)
            new_first_name = random.choice(first_names)
            vietnamese_names.append((new_last_name, new_middle_name, new_first_name))

        for _ in range(num_australian):
            new_first_name, new_last_name = random.choice(AUSTRALIAN_NAMES)
            australian_names.append((new_last_name, "", new_first_name))

        for _ in range(num_foreign):
            new_last_name, new_first_name = random.choice(FOREIGN_NAMES)
            foreign_names.append((new_last_name, "", new_first_name))

        all_names = vietnamese_names + australian_names + foreign_names
        random.shuffle(all_names)

        fields = ['STT', 'Họ và tên', 'Hãng HK', 'Số ghế ngồi']
        new_data = pd.DataFrame(columns=fields)
        new_data['STT'] = range(1, num_new_names + 1)
        new_data['Họ và tên'] = [' '.join(names) for names in all_names]
        new_data['Số ghế ngồi'] = self.generate_seat_numbers(num_new_names)
        new_data['Hãng HK'] = [airline_name] * num_new_names

        return new_data, fields

    def post(self):
        try:
            FOREIGN_NAMES = [
                ('Smith', 'John'), ('Johnson', 'Robert'), ('Williams', 'Michael'), ('Jones', 'David'), ('Brown', 'William'),
                ('Davis', 'Richard'), ('Miller', 'Joseph'), ('Wilson', 'Charles'), ('Moore', 'Thomas'), ('Taylor', 'Daniel'),
                ('Thomas', 'Matthew'), ('Harris', 'Donald'), ('Clark', 'Anthony'), ('Lewis', 'Mark'), ('Lee', 'Paul'),
                ('Walker', 'Steven'), ('Hall', 'Kevin'), ('Allen', 'Edward'), ('Young', 'Brian'), ('King', 'Ronald'),
                ('White', 'George'), ('Green', 'Kenneth'), ('Turner', 'Andrew'), ('Cook', 'Jeffrey'), ('Baker', 'Timothy'),
                ('Hill', 'Steven'), ('Carter', 'Jerry'), ('Roberts', 'Frank'), ('Wood', 'Scott'), ('Wright', 'Christopher')
            ]
            AUSTRALIAN_NAMES = [
                ('Allan', 'Rebelo'), ('Bianca', 'Garofalo'), ('Brandon', 'Parson'), ('Brittany', 'Hastings'),
                ('Cheyenne', 'Jurcik'), ('Colby', 'Taylor'), ('Eddy', 'Quispillo'), ('Franco', 'Masdea'),
                ('Haley', 'Anderson'), ('Jacqueline', 'Rose'), ('Jared', 'David'), ('Josh', 'Buswa'),
                ('Jason', 'Smith'), ('Evie', 'Taylor'), ('Oliver', 'Campbell'), ('Noah', 'Johnson'),
                ('William', 'Walker'), ('Henry', 'Graham'), ('Andy', 'Turnbel'), ('Olivia', 'Lee'),
                ('Mia', 'Robinson'), ('Charlotte', 'Scott'), ('Amelia', 'Sanders'), ('Isla', 'Ball')
            ]
            file = request.files['file']
            airline_name = request.form.get('airline')
            date = request.form.get('date')
            fromm = request.form.get('from')
            to = request.form.get('to')
            time = request.form.get('time')
            start = request.form.get('start')
            end = request.form.get('end')
            fill_color = self.convert_color_to_argb_hex(request.form.get('fill-color'))
            font_color = self.convert_color_to_argb_hex(request.form.get('font-color'))
            num_new_names = min(max(int(request.form.get('quantity')), 1), 315)
            font_family = request.form.get('font-family') or "Times New Roman"
            size_input = request.form.get('size')
            size = min(max(int(size_input), 10), 32) if size_input else 10

            data = pd.read_excel(file)

            last_names = data['Họ'].tolist()
            middle_names = data['Đệm'].apply(lambda x: str(x).lstrip() if not pd.isna(x) else "").tolist()
            first_names = data['Tên'].tolist()

            new_data, fields = self.generate_data(num_new_names, last_names, middle_names, first_names, airline_name, AUSTRALIAN_NAMES, FOREIGN_NAMES)

            wb = Workbook()
            sheet = wb.active
            sheet.title = 'Danh sách hành khách'

            sheet.column_dimensions['A'].width = 10
            sheet.column_dimensions['B'].width = size*2.5
            sheet.column_dimensions['C'].width = size*2.5
            sheet.column_dimensions['D'].width = size*1.5

            sheet.row_dimensions[1].height = size + 10 + 4
            sheet.row_dimensions[2].height = size + 10 + 4
            sheet.row_dimensions[3].height = size + 4 + 2

            alignment = Alignment(horizontal='center', vertical='center')
            alignment2 = Alignment(horizontal='left', vertical='center')

            font = Font(name=font_family, size=size+3, color=font_color, bold=True)
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            dateobj =  datetime.datetime.strptime(date, "%Y-%m-%d").date()
            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
            sheet['A1'] = airline_name + " "*5 + str(dateobj.strftime('%A')) + " " + str(dateobj)
            sheet['A1'].font = font
            sheet['A1'].fill = fill
            sheet['A1'].alignment = alignment2
            sheet['A1'].border = Border(left=Side(style='thin'), top=Side(style='thin'))
            sheet['B1'].border = Border(top=Side(style='thin'))


            sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
            sheet['A2'] = fromm + " - " + to
            sheet['A2'].font = font
            sheet['A2'].fill = fill
            sheet['A2'].alignment = alignment2
            sheet['A2'].border = Border(left=Side(style='thin'))

            sheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=3)
            sheet['C1'] = "Thời gian bay: " + str(time) + "h"
            sheet['C1'].font = font
            sheet['C1'].fill = fill
            sheet['C1'].alignment = alignment2
            sheet['C1'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'))


            sheet.merge_cells(start_row=2, start_column=3, end_row=2, end_column=3)
            sheet['C2'] = "Từ: " + str(start) + " - " + str(end)
            sheet['C2'].font = font
            sheet['C2'].fill = fill
            sheet['C2'].alignment = alignment2
            sheet['C2'].border = Border(left=Side(style='thin'), right=Side(style='thin'))

            sheet.merge_cells(start_row=1, start_column=4, end_row=2, end_column=4)
            sheet['D1'] = "Tổng số ghế: " + str(num_new_names)
            sheet['D1'].font = font
            sheet['D1'].fill = fill
            sheet['D1'].alignment = alignment
            sheet['D1'].border = Border(right=Side(style='thin'), top=Side(style='thin'))
            sheet['D2'].border = Border(right=Side(style='thin'))

            sheet.append(fields)
            field_font = Font(name=font_family, size=size+2, bold=True)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            sheet.merge_cells(start_row=3, start_column=2, end_row=3, end_column=3)
            for col_index, col_name in enumerate(fields, start=1):
                cell = sheet.cell(row=3, column=col_index)
                cell.font = field_font
                cell.border = border
                cell.alignment = alignment
                cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

            even_row_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
            odd_row_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            value_font = Font(name=font_family, size=size)
            
            current_row = 4
            for _, row in new_data.iterrows():
                row_values = [row.STT, row['Họ và tên'], '', row['Số ghế ngồi']]
                sheet.append(row_values)
                sheet.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
                for cell in sheet[current_row]:
                    cell.border = border
                    cell.font = value_font
                    if cell.column != 2:
                        cell.alignment = alignment
                    cell.fill = even_row_fill if current_row % 2 == 0 else odd_row_fill
                current_row += 1

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            return send_file(output, download_name='generated_data.xlsx', as_attachment=True)

        except Exception as e:
            return f"An error occurred: {str(e)}"

app.add_url_rule('/upload', view_func=ExcelGeneratorView.as_view('upload'))
    
if __name__ == '__main__':
    app.run()
